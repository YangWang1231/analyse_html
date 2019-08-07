#!/usr/bin/python
 #coding:utf-8

'''
构建数据库中的LDRA_rule table
read from LDRA973-support-Gjb8114.xlsx, sheet 'GJB8114-rule-details', get data of columes
and insert data into three tables:
1.GBJ8114_rule
2.LDRA_rule
3.GJB_LDRA_relation_table

PS.程序中应该有一个地方保存全局设置，可以参考flask

'''

from openpyxl import load_workbook
from store_db_sqlit3 import process_db
import re
from config import get_config
#a lot of repeat pattern, etc open db, load workbook, read sheet, iterate sheet, process row data
#so may be a decorated pattern should be used.

#fetch GJB8114 explainsion-chs from file gjb8114 rule in chinese.xlsx



def fill_8114_chineses():
    config_data = get_config()
    db_path = config_data['DB_location'] + config_data['DB_name']
    db_obj = process_db(db_path)

   #db_obj.init_db()
    rule_translate_chines = config_data["gjb8114 rule in chinese"]['filename']
    translate_sheet = config_data["gjb8114 rule in chinese"]['sheetname']
    wb = load_workbook(rule_translate_chines)
    ws = wb[translate_sheet]
    with open('chinese-rule.dat','w') as f:
        rule_chinese_dic = {}
        for row in ws.values:
            rule_ches_string = row[0].encode('utf-8') #encode unicode to utf-8, so re can process it
            #split rule-chinese pair, regex pattern
            match_rule_mandatory_string = u'[RA]-\d-\d+-\d+' #匹配规则
            
            regex_rule = re.compile(match_rule_mandatory_string)
            for e in regex_rule.finditer(rule_ches_string):
                rule, chinese = rule_ches_string[e.start() : e.end()].strip() , rule_ches_string[e.end() : ].strip()
                rule_chinese_dic[rule] = chinese
                print(rule, chinese)
                f.write('%s\n' %chinese)
                f.flush()
                db_obj.update_GJB8114_chinese((chinese.decode('utf-8'), rule)) #encode utf-8 to unicode, so data can put inot DB
    #fill DB
    db_obj.commit()
    
    return

#表格LDRA973-support-Gjb8114.xlsx的GJB8114-rule-details保存了对应关系
rule_xlsx_file_name = 'LDRA973-support-Gjb8114.xlsx'
rule_sheet_name = u'GJB8114-rule-details'
def build_LDRA_rule_table():
    """from LDRA973-support-Gjb8114.xlsx import information, and stroe to db
    fill GJB_8114_rule, LDRA_rul, GJB_LDRA_relation_table
    """
#    db_obj = process_db()
    config_data = get_config()
    db_path = config_data['DB_location'] + config_data['DB_name']
    db_obj = process_db(db_path)

    rule_xlsx_file_name = config_data['LDRA-support-Gjb8114']['filename']
    rule_sheet_name = config_data['LDRA-support-Gjb8114']['sheetname']
    wb = load_workbook(rule_xlsx_file_name)
    ws = wb[rule_sheet_name]

    last_ldra_code = ''
    for row in ws.values:
        print(row)
        GJB8114Code, LDRACode, MandatoryStanard_LDRA, MandatoryStandard_8114, Rule_classification = row
        #if GJB8114Code == None:     #处理多个testbedrule共同支持一条gjb8114规则的情况
        #    #这本来就说明，对应关系决定了应该拆分成两张表格 
        #    pass

        #ldracode存在与GJB的对应，并且不重复
        if LDRACode != '' :
            ldraid = db_obj.get_ldra_id(LDRACode)  
            if ldraid == None:
                ldraid = db_obj.insert_LDRA_rule((LDRACode, MandatoryStanard_LDRA))            
        
        #gjb不重复
        if GJB8114Code != None:
            Rule_classification = Rule_classification.upper()
            gjb_id = db_obj.insert_GJB8114_rule((GJB8114Code, MandatoryStandard_8114, '', Rule_classification))

            #当GJB中存在一个条目，且LDRACode也存在，就需要在relation表格中建立一个新的条目
        if LDRACode != '' :
            db_obj.insert_GJB_LDRA_relation_table((gjb_id, ldraid))
        #save last item
        #last_GJB8114_code = GJB8114Code
        #last_MandatoryStandard_ch = MandatoryStandard_8114
        #last_Rule_classification = Rule_classification

    db_obj.commit()


if __name__ == '__main__':
    #fill_8114_chineses()
    build_LDRA_rule_table()
